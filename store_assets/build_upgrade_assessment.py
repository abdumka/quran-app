#!/usr/bin/env python3
"""
build_upgrade_assessment.py   v1.0
OHSU Network Engineering — Fabric Upgrade Assessment Workbook Builder

Recreates the DCW_upgrade_assessment.xlsx format for any blueprint (DCB, BIC05, etc.)

USAGE
-----
  1. Create a device list CSV (hostname,mgmt_ip) — one line per switch:

       DCBbl111a,10.86.243.30
       DCBbl111b,10.86.243.31
       DCBlf10102a,10.86.243.40
       DCBlf10102b,10.86.243.41
       ...

     Pairs are detected automatically by stripping the trailing a/b.

  2. Collect + build in one shot:
       python3 build_upgrade_assessment.py --devices dcb_devices.csv --blueprint DCB

  3. Or collect once, rebuild the workbook offline as many times as you like:
       python3 build_upgrade_assessment.py --devices dcb_devices.csv --blueprint DCB --collect-only
       python3 build_upgrade_assessment.py --from-cache DCB_raw.json --blueprint DCB

Password is prompted securely. Domain usernames work: --user "ohsum01\\keilani"
"""

import argparse
import csv
import datetime
import getpass
import json
import os
import re
import sys
from collections import OrderedDict, defaultdict

try:
    import paramiko
except ImportError:
    sys.exit("ERROR: paramiko not installed.  pip install paramiko")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl not installed.  pip install openpyxl")


# ─────────────────────────────────────────────────────────────────────────────
# Commands collected from every switch
# ─────────────────────────────────────────────────────────────────────────────
COMMANDS = OrderedDict([
    ("version",     "show version"),
    ("uptime",      "show system uptime"),
    ("alarms",      "show chassis alarms"),
    ("intf_desc",   "show interfaces descriptions"),
    ("intf_config", "show configuration interfaces | display set"),
    ("mac_table",   "show ethernet-switching table"),
    ("mac_table_alt", "show mac-table"),
])

# ─────────────────────────────────────────────────────────────────────────────
# Styling constants (match the DCW workbook)
# ─────────────────────────────────────────────────────────────────────────────
FONT = "Arial"
C_HDR_BLUE   = "4472C4"   # auto-filled header
C_HDR_YELLOW = "FFF2CC"   # manual-entry header
C_ORANGE     = "FFD966"   # dual-connected, no LAG
C_RED        = "F4B6B6"   # single-homed
C_BLUE_ROW   = "BDD7EE"   # split dual-path / assumed
C_GREEN      = "E2EFDA"   # healthy
C_TITLE      = "1F3864"
C_BAND       = "F2F2F2"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(hex_):
    return PatternFill("solid", fgColor=hex_)


def hdr_cell(ws, row, col, text, bg=C_HDR_BLUE, white=True):
    c = ws.cell(row, col, text)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF" if white else "000000")
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    return c


def data_cell(ws, row, col, text, bg=None, wrap=False, size=9):
    c = ws.cell(row, col, text)
    c.font = Font(name=FONT, size=size)
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(vertical="top", wrap_text=wrap)
    c.border = BORDER
    return c


# ─────────────────────────────────────────────────────────────────────────────
# SSH collection
# ─────────────────────────────────────────────────────────────────────────────
class Switch:
    def __init__(self, host, ip, user, password):
        self.host = host
        self.ip = ip
        self.user = user
        self.password = password
        self.client = None
        self.raw = {}

    def connect(self, timeout=25):
        self.client = paramiko.SSHClient()
        self.client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        self.client.connect(
            hostname=self.ip, username=self.user, password=self.password,
            timeout=timeout, look_for_keys=False, allow_agent=False,
        )

    def run(self, cmd, timeout=90):
        try:
            _, out, err = self.client.exec_command(f"{cmd} | no-more", timeout=timeout)
            return out.read().decode("utf-8", "replace") or err.read().decode("utf-8", "replace")
        except Exception as e:
            return f"[ERROR: {e}]"

    def collect(self):
        for key, cmd in COMMANDS.items():
            self.raw[key] = self.run(cmd)
        # pick whichever MAC command actually worked
        if "syntax error" in self.raw.get("mac_table", "").lower() or not self.raw.get("mac_table", "").strip():
            self.raw["mac_table"] = self.raw.get("mac_table_alt", "")
        self.raw.pop("mac_table_alt", None)
        return self.raw

    def close(self):
        if self.client:
            try:
                self.client.close()
            except Exception:
                pass


# ─────────────────────────────────────────────────────────────────────────────
# Parsers
# ─────────────────────────────────────────────────────────────────────────────
def parse_version(text):
    model = re.search(r"Model:\s+(\S+)", text)
    junos = re.search(r"Junos:\s+(\S+)", text)
    return (model.group(1) if model else "unknown",
            junos.group(1) if junos else "unknown")


def parse_uptime(text):
    m = re.search(r"System booted:.*?\((.*?)\s+ago\)", text)
    if m:
        return m.group(1)
    m = re.search(r"up\s+([^\n,]+)", text)
    return m.group(1).strip() if m else "unknown"


def parse_alarms(text):
    lines = [l.rstrip() for l in text.splitlines() if l.strip()]
    count = 0
    m = re.search(r"(\d+)\s+alarms?\s+currently active", text)
    if m:
        count = int(m.group(1))
    return count, lines


def parse_intf_desc(text):
    """show interfaces descriptions -> {intf: (admin, link, description)}"""
    out = OrderedDict()
    for line in text.splitlines():
        m = re.match(r"^(\S+)\s+(up|down)\s+(up|down)\s*(.*)$", line.strip())
        if m:
            intf, admin, link, desc = m.groups()
            if intf.lower() == "interface":
                continue
            out[intf] = (admin, link, desc.strip())
    return out


def parse_intf_config(text):
    """
    show configuration interfaces | display set
    Returns:
      ae_members  {ae: [member,...]}
      member_ae   {member: ae}
      esi         {ae: value}
      esi_mode    {ae: mode}
      lacp        {ae: bool}
      descs       {intf: description}
    """
    ae_members = defaultdict(list)
    member_ae = {}
    esi, esi_mode, lacp, descs = {}, {}, {}, {}

    for line in text.splitlines():
        line = line.strip()
        m = re.match(r"set interfaces (\S+) ether-options 802\.3ad (\S+)", line)
        if m:
            member, ae = m.groups()
            ae_members[ae].append(member)
            member_ae[member] = ae
            continue
        m = re.match(r"set interfaces (\S+) esi (\S+:\S+)", line)
        if m:
            esi[m.group(1)] = m.group(2)
            continue
        m = re.match(r"set interfaces (\S+) esi (all-active|single-active)", line)
        if m:
            esi_mode[m.group(1)] = m.group(2)
            continue
        if "aggregated-ether-options lacp active" in line:
            m = re.match(r"set interfaces (\S+) ", line)
            if m:
                lacp[m.group(1)] = True
            continue
        m = re.match(r'set interfaces (\S+) description "?([^"]+)"?', line)
        if m:
            descs[m.group(1)] = m.group(2).strip()

    return dict(ae_members), member_ae, esi, esi_mode, lacp, descs


def parse_mac_table(text):
    """Returns list of (vlan, mac, flags, interface)."""
    rows = []
    for line in text.splitlines():
        m = re.match(
            r"^\s*(\S+)\s+([0-9a-fA-F:]{17})\s+(\S+)\s+(\S+)", line)
        if m:
            rows.append(m.groups())
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Host redundancy analysis
# ─────────────────────────────────────────────────────────────────────────────
UPLINK_RE = re.compile(r"facing_|to\.\w*sp\d|spine", re.I)


def analyse_pair(a_name, a_data, b_name, b_data):
    """
    Group interfaces by description across both switches and classify.
    Returns (hosts, risk_flags) where hosts is a list of dicts.
    """
    def side(data):
        desc = parse_intf_desc(data["intf_desc"])
        ae_members, member_ae, esi, esi_mode, lacp, _ = parse_intf_config(data["intf_config"])
        return desc, ae_members, member_ae, esi, esi_mode, lacp

    a_desc, a_aem, a_mae, a_esi, a_mode, a_lacp = side(a_data)
    b_desc, b_aem, b_mae, b_esi, b_mode, b_lacp = side(b_data) if b_data else ({}, {}, {}, {}, {}, {})

    # group physical interfaces by description, skipping fabric uplinks and AEs
    def group(desc_map):
        g = defaultdict(list)
        for intf, (admin, link, d) in desc_map.items():
            if not d or "." in intf or intf.startswith("ae") or intf.startswith("lo"):
                continue
            if UPLINK_RE.search(d):
                continue
            g[d].append(intf)
        return g

    a_hosts, b_hosts = group(a_desc), group(b_desc)
    all_hosts = sorted(set(a_hosts) | set(b_hosts))

    rows, flags = [], []

    for host in all_hosts:
        a_ifs = sorted(a_hosts.get(host, []))
        b_ifs = sorted(b_hosts.get(host, []))

        ae = next((a_mae[i] for i in a_ifs if i in a_mae), None) \
             or next((b_mae[i] for i in b_ifs if i in b_mae), None)

        esi_val = a_esi.get(ae) or b_esi.get(ae) if ae else None
        mode = a_mode.get(ae) or b_mode.get(ae) if ae else None
        has_lacp = bool(a_lacp.get(ae) or b_lacp.get(ae)) if ae else False

        host_flags = []

        if a_ifs and b_ifs:
            if ae and esi_val:
                conn = "Dual-homed (ESI-LAG)"
            elif ae:
                conn = "Dual-homed (LACP)"
                host_flags.append(f"AE {ae} is missing ESI")
            else:
                conn = "Dual-connected (No LAG)"
                host_flags.append(
                    "Dual-connected without LAG/ESI — each NIC is independent. "
                    "Verify host NIC bonding/teaming before upgrading either switch.")
        else:
            conn = "Single-homed"
            only = a_name if a_ifs else b_name
            host_flags.append(f"Single-homed on {only} — AT RISK during that switch's upgrade")

        # admin-up / link-down detection
        for intf in a_ifs:
            admin, link, _ = a_desc[intf]
            if admin == "up" and link == "down":
                host_flags.append(f"{intf} is admin-up but link-down on switch A")
        for intf in b_ifs:
            admin, link, _ = b_desc[intf]
            if admin == "up" and link == "down":
                host_flags.append(f"{intf} is admin-up but link-down on switch B")

        rows.append({
            "host": host,
            "a_ifs": ", ".join(a_ifs) or "—",
            "b_ifs": ", ".join(b_ifs) or "—",
            "ae": ae or "—",
            "esi": esi_val or ("MISSING" if ae else "—"),
            "esi_mode": mode or "—",
            "lacp": "Yes" if has_lacp else "No",
            "conn": conn,
            "flags": " | ".join(host_flags) if host_flags else "OK",
        })
        flags.extend(host_flags)

    return rows, flags


# ─────────────────────────────────────────────────────────────────────────────
# Workbook builder
# ─────────────────────────────────────────────────────────────────────────────
def build_workbook(blueprint, devices, cache, outfile):
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "SUMMARY"

    # pair devices by stripping trailing a/b
    pairs = defaultdict(dict)
    for host, ip in devices:
        m = re.match(r"^(.*?)([ab])$", host, re.I)
        if m:
            pairs[m.group(1).lower()][m.group(2).lower()] = host
        else:
            pairs[host.lower()]["a"] = host

    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── SUMMARY header ──
    t = ws_sum.cell(1, 1, f"  Upgrade Assessment — Blueprint: {blueprint}")
    t.font = Font(name=FONT, size=14, bold=True, color=C_TITLE)
    ws_sum.cell(2, 1, f"Generated: {ts}").font = Font(name=FONT, size=9, italic=True, color="595959")

    sum_hdrs = ["Pair ID", "Switch A", "Switch B", "Total Hosts", "Dual (ESI/LAG)",
                "No-LAG Dual", "Single-homed", "Active Alarms",
                "Risk Flags (summary)", "Safe to Upgrade?", "Upgrade Date"]
    for i, h in enumerate(sum_hdrs, 1):
        hdr_cell(ws_sum, 4, i, h)

    all_hosts_rows = []
    review_rows = []
    srow = 5

    for pair_id in sorted(pairs):
        a_name = pairs[pair_id].get("a")
        b_name = pairs[pair_id].get("b")
        a_data = cache.get(a_name)
        b_data = cache.get(b_name) if b_name else None
        if not a_data:
            continue

        rows, flags = analyse_pair(a_name, a_data, b_name, b_data)

        n_dual = sum(1 for r in rows if r["conn"].startswith("Dual-homed"))
        n_nolag = sum(1 for r in rows if r["conn"] == "Dual-connected (No LAG)")
        n_single = sum(1 for r in rows if r["conn"] == "Single-homed")

        alarm_ct = 0
        for d in (a_data, b_data):
            if d:
                c, _ = parse_alarms(d["alarms"])
                alarm_ct += c

        flag_txt = " | ".join(flags[:3])
        if len(flags) > 3:
            flag_txt += " ..."

        bg = C_BAND if (srow % 2 == 0) else None
        vals = [pair_id, a_name, b_name or "—", len(rows), n_dual, n_nolag,
                n_single, alarm_ct, flag_txt, "", ""]
        for i, v in enumerate(vals, 1):
            data_cell(ws_sum, srow, i, v, bg=bg, wrap=(i == 9))
        # highlight Safe-to-Upgrade / Date as user-entry cells
        for i in (10, 11):
            ws_sum.cell(srow, i).fill = fill(C_HDR_YELLOW)
        srow += 1

        # ── per-switch detail sheets ──
        for label, name, data in (("A", a_name, a_data), ("B", b_name, b_data)):
            if not data or not name:
                continue
            build_switch_sheet(wb, blueprint, name, label, data)

        # ── accumulate combined + review rows ──
        first = True
        for r in rows:
            all_hosts_rows.append([
                f"{a_name} / {b_name}" if first else None,
                r["host"], r["a_ifs"], r["b_ifs"], r["conn"],
                None, None, None, None, None, None,
            ])
            review_rows.append([
                a_name, b_name or "—", r["host"], r["a_ifs"], r["b_ifs"],
                r["conn"], None, None, None, None, None, None, None, None,
            ])
            first = False

    ws_sum.freeze_panes = "A5"
    for i, w in enumerate([14, 16, 16, 11, 14, 12, 13, 13, 60, 20, 16], 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w

    build_combined(wb, all_hosts_rows)
    build_cab_review(wb, all_hosts_rows)
    build_hosts_for_review(wb, blueprint, ts, review_rows)

    wb.save(outfile)
    return outfile


def build_switch_sheet(wb, blueprint, name, label, data):
    ws = wb.create_sheet(name[:31])
    model, junos = parse_version(data["version"])
    up = parse_uptime(data["uptime"])

    t = ws.cell(1, 1, f"  {name}  ({label} Switch)  —  Blueprint: {blueprint}")
    t.font = Font(name=FONT, size=13, bold=True, color=C_TITLE)

    ws.cell(2, 1, "Management IP:").font = Font(name=FONT, size=9, bold=True)
    ws.cell(2, 2, data.get("_ip", "")).font = Font(name=FONT, size=9)
    ws.cell(2, 4, "OS / Profile:").font = Font(name=FONT, size=9, bold=True)
    ws.cell(2, 5, f"Junos {junos}  {model}").font = Font(name=FONT, size=9)
    ws.cell(3, 1, "Junos Version:").font = Font(name=FONT, size=9, bold=True)
    ws.cell(3, 2, f"{model}  |  Junos {junos}").font = Font(name=FONT, size=9)
    ws.cell(3, 4, "Uptime:").font = Font(name=FONT, size=9, bold=True)
    ws.cell(3, 5, up).font = Font(name=FONT, size=9)

    r = 5
    ws.cell(r, 1, "CHASSIS ALARMS").font = Font(name=FONT, size=11, bold=True, color=C_TITLE)
    r += 1
    _, alines = parse_alarms(data["alarms"])
    for line in alines:
        ws.cell(r, 1, line).font = Font(name="Courier New", size=8)
        r += 1
    r += 1

    # interface details
    ws.cell(r, 1, "INTERFACE DETAILS").font = Font(name=FONT, size=11, bold=True, color=C_TITLE)
    r += 1
    for i, h in enumerate(["Interface", "Admin", "Link", "Description", "AE Member", "Notes"], 1):
        hdr_cell(ws, r, i, h)
    r += 1
    ae_members, member_ae, esi, esi_mode, lacp, _ = parse_intf_config(data["intf_config"])
    for intf, (admin, link, desc) in parse_intf_desc(data["intf_desc"]).items():
        note = "Admin-up but link-down" if (admin == "up" and link == "down") else None
        data_cell(ws, r, 1, intf)
        data_cell(ws, r, 2, admin)
        data_cell(ws, r, 3, link)
        data_cell(ws, r, 4, desc)
        data_cell(ws, r, 5, member_ae.get(intf))
        data_cell(ws, r, 6, note, bg=C_ORANGE if note else None)
        r += 1
    r += 1

    # AE / ESI
    ws.cell(r, 1, "AE / ESI-LAG INTERFACES").font = Font(name=FONT, size=11, bold=True, color=C_TITLE)
    r += 1
    for i, h in enumerate(["AE Interface", "ESI Value", "ESI Mode", "LACP Active", "Member Interfaces"], 1):
        hdr_cell(ws, r, i, h)
    r += 1
    for ae in sorted(ae_members):
        val = esi.get(ae, "NOT CONFIGURED")
        data_cell(ws, r, 1, ae)
        data_cell(ws, r, 2, val, bg=None if ae in esi else C_ORANGE)
        data_cell(ws, r, 3, esi_mode.get(ae, "—"))
        data_cell(ws, r, 4, "Yes" if lacp.get(ae) else "No")
        data_cell(ws, r, 5, ", ".join(sorted(ae_members[ae])))
        r += 1
    r += 1

    # MAC table
    macs = parse_mac_table(data.get("mac_table", ""))
    if macs:
        ws.cell(r, 1, "MAC ADDRESS TABLE").font = Font(name=FONT, size=11, bold=True, color=C_TITLE)
        r += 1
        for i, h in enumerate(["VLAN", "MAC Address", "Flags", "Interface"], 1):
            hdr_cell(ws, r, i, h)
        r += 1
        for vlan, mac, flg, intf in macs:
            data_cell(ws, r, 1, vlan)
            data_cell(ws, r, 2, mac)
            data_cell(ws, r, 3, flg)
            data_cell(ws, r, 4, intf)
            r += 1

    for i, w in enumerate([22, 14, 12, 42, 14, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_combined(wb, rows):
    ws = wb.create_sheet("All Hosts - Combined")
    hdrs = ["Switch Pair", "Host / Device", "Port — Switch A", "Port — Switch B",
            "Connection Type", "Internal Redundancy?", "Host Owner",
            "OK to Bring Down?", "Host Importance", "Upgrade Status", "Upgrade Date"]
    for i, h in enumerate(hdrs, 1):
        hdr_cell(ws, 1, i, h)
    for r, row in enumerate(rows, 2):
        for i, v in enumerate(row, 1):
            bg = None
            if i == 5 and v:
                bg = {"Dual-connected (No LAG)": C_ORANGE,
                      "Single-homed": C_RED}.get(v)
            data_cell(ws, r, i, v, bg=bg)
        for i in range(6, 12):
            ws.cell(r, i).fill = fill(C_HDR_YELLOW)
    ws.freeze_panes = "A2"
    for i, w in enumerate([26, 34, 22, 22, 24, 18, 14, 18, 16, 18, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_cab_review(wb, rows):
    """Only rows that need human attention: no-LAG dual and single-homed."""
    ws = wb.create_sheet("CAB Review")
    hdrs = ["Pair ID", "Host / Device", "Port — Sw A", "Port — Sw B", "Connection Type",
            "Internal Redundancy?", "Host Owner", "OK to Bring Down?",
            "Upgrade Status", "Upgrade Date"]
    for i, h in enumerate(hdrs, 1):
        hdr_cell(ws, 1, i, h)
    r = 2
    current_pair = None
    last_written = None
    for row in rows:
        if row[0]:
            current_pair = row[0]
        conn = row[4]
        if conn not in ("Dual-connected (No LAG)", "Single-homed"):
            continue
        pair_label = current_pair if current_pair != last_written else None
        if pair_label:
            last_written = current_pair
        vals = [pair_label, row[1], row[2], row[3], conn, None, None, None, None, None]
        for i, v in enumerate(vals, 1):
            bg = C_ORANGE if (i == 5 and conn == "Dual-connected (No LAG)") else \
                 C_RED if (i == 5) else None
            data_cell(ws, r, i, v, bg=bg)
        for i in range(6, 11):
            ws.cell(r, i).fill = fill(C_HDR_YELLOW)
        r += 1
    ws.freeze_panes = "A2"
    for i, w in enumerate([16, 34, 18, 18, 24, 18, 14, 20, 20, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_hosts_for_review(wb, blueprint, ts, rows):
    ws = wb.create_sheet("Hosts for Review")
    t = ws.cell(1, 1, f"  {blueprint}  —  Server / Host Review Sheet     Generated: {ts}")
    t.font = Font(name=FONT, size=13, bold=True, color=C_TITLE)
    legend = ("Auto-filled columns (blue) come from the switch assessment. "
              "Yellow columns must be completed by the server team / host owner. "
              "Connection Status: ORANGE = dual-connected but no LAG (verify host-side bonding);  "
              "RED = single-homed (at risk during upgrade);  "
              "BLUE = script assumed split dual-path (verify MPIO/multipath — see Notes column).")
    c = ws.cell(2, 1, legend)
    c.font = Font(name=FONT, size=8, italic=True, color="595959")

    hdrs = ["Switch A", "Switch B", "Host / Description", "Interface(s)\non Switch A",
            "Interface(s)\non Switch B", "Connection Status", "Host Owner",
            "Host Type /\nPurpose", "Internal\nRedundancy?", "Redundancy\nMechanism",
            "OK to Bring Down one side\nDuring Upgrade?", "Host\nImportance",
            "Needs Separate\nDowntime?", "Notes / Action Items"]
    for i, h in enumerate(hdrs, 1):
        hdr_cell(ws, 3, i, h, bg=C_HDR_BLUE if i <= 6 else C_HDR_YELLOW, white=(i <= 6))

    r = 4
    for row in rows:
        conn = row[5]
        if conn == "Dual-homed (ESI-LAG)":
            continue          # fully redundant — no host-owner input needed
        for i, v in enumerate(row, 1):
            bg = None
            if i == 6:
                bg = {"Dual-connected (No LAG)": C_ORANGE,
                      "Single-homed": C_RED,
                      "Dual-homed (LACP)": C_BLUE_ROW}.get(conn)
            elif i >= 7:
                bg = C_HDR_YELLOW
            data_cell(ws, r, i, v, bg=bg, wrap=(i in (3, 14)))
        r += 1

    ws.freeze_panes = "A4"
    for i, w in enumerate([16, 16, 32, 18, 18, 24, 14, 16, 14, 18, 24, 14, 16, 34], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="Fabric upgrade assessment workbook builder")
    ap.add_argument("--devices", help="CSV: hostname,mgmt_ip (one per line)")
    ap.add_argument("--blueprint", required=True, help="Blueprint name, e.g. DCB")
    ap.add_argument("--user", default="keilani", help="SSH username")
    ap.add_argument("--password", default=None, help="SSH password (prompts if omitted)")
    ap.add_argument("--collect-only", action="store_true", help="Collect raw data, skip workbook")
    ap.add_argument("--from-cache", help="Build workbook from an existing raw JSON file")
    ap.add_argument("--out", help="Output .xlsx filename")
    args = ap.parse_args()

    cache_file = args.from_cache or f"{args.blueprint}_raw.json"
    outfile = args.out or f"{args.blueprint}_upgrade_assessment.xlsx"

    devices = []
    if args.devices:
        with open(args.devices, newline="") as f:
            for row in csv.reader(f):
                if not row or row[0].strip().startswith("#"):
                    continue
                if len(row) < 2:
                    print(f"  [skip] malformed line: {row}")
                    continue
                devices.append((row[0].strip(), row[1].strip()))

    # ── build from cache ──
    if args.from_cache:
        with open(cache_file) as f:
            cache = json.load(f)
        if not devices:
            devices = [(h, cache[h].get("_ip", "")) for h in cache]
        print(f"Building workbook from {cache_file} ...")
        out = build_workbook(args.blueprint, devices, cache, outfile)
        print(f"[SAVED] {out}")
        return

    if not devices:
        sys.exit("ERROR: --devices is required unless --from-cache is used")

    password = args.password or getpass.getpass(f"Password for {args.user}: ")

    cache = {}
    print(f"\nCollecting from {len(devices)} device(s)...\n")
    for host, ip in devices:
        print(f"  [{host}]  {ip}")
        sw = Switch(host, ip, args.user, password)
        try:
            sw.connect()
        except Exception as e:
            print(f"      !! connect failed: {e}")
            continue
        raw = sw.collect()
        raw["_ip"] = ip
        cache[host] = raw
        sw.close()

    with open(cache_file, "w") as f:
        json.dump(cache, f, indent=1)
    print(f"\n[SAVED] raw data: {cache_file}")

    if args.collect_only:
        print("Collect-only mode — skipping workbook.")
        return

    out = build_workbook(args.blueprint, devices, cache, outfile)
    print(f"[SAVED] {out}")


if __name__ == "__main__":
    main()
